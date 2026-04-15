"""Data preprocessing and template validation for jamovi-analysis."""

import csv
import json
import re
import tempfile
from pathlib import Path
from typing import Any


class PreprocessError(Exception):
    """Raised when data preprocessing or template validation fails."""


def safe_alias(name: str, index: int, seen: set[str]) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", str(name).strip())
    cleaned = re.sub(r"__+", "_", cleaned).strip("_")
    if not cleaned or not cleaned[0].isalpha():
        cleaned = f"var_{cleaned}" if cleaned else f"var_{index}"
    base = cleaned
    counter = 1
    while cleaned in seen:
        cleaned = f"{base}_{counter}"
        counter += 1
    seen.add(cleaned)
    return cleaned


def _validate_template(
    dict_rows: list[dict[str, Any]],
    aliases: list[str],
    manifest: dict[str, str],
    spec: dict[str, Any],
) -> None:
    """Validate data against template_hint constraints."""
    template_hint = spec.get("template_hint")
    if template_hint is None:
        return

    def _resolve_item(item_name: str) -> str | None:
        if item_name in manifest:
            return item_name
        for k, v in manifest.items():
            if v == item_name:
                return k
        return None

    if template_hint == "reliability_scale":
        min_scale = spec.get("min_scale", 1)
        max_scale = spec.get("max_scale", 5)
        scale_aliases = [a for a in aliases if re.match(r"^q\d+$", a, re.IGNORECASE)]
        if not scale_aliases:
            raise PreprocessError(
                "Template 'reliability_scale' requires at least one column matching 'q01', 'q02', etc."
            )
        for alias in scale_aliases:
            for i, row in enumerate(dict_rows):
                val = row.get(alias)
                if val is None or val == "":
                    continue
                try:
                    num = float(val)
                    if not num.is_integer():
                        raise PreprocessError(
                            f"Column '{manifest[alias]}' (row {i + 1}) contains a non-integer value: {val}. "
                            f"Reliability scale data must be integers between {min_scale} and {max_scale}."
                        )
                    num = int(num)
                    if num < min_scale or num > max_scale:
                        raise PreprocessError(
                            f"Column '{manifest[alias]}' (row {i + 1}) value {num} is out of range. "
                            f"Expected integers between {min_scale} and {max_scale}."
                        )
                except (ValueError, TypeError):
                    raise PreprocessError(
                        f"Column '{manifest[alias]}' (row {i + 1}) contains an invalid value: {val}. "
                        f"Reliability scale data must be integers between {min_scale} and {max_scale}."
                    )

    elif template_hint == "ttest_two_group":
        group_col = spec.get("group_column", "group")
        group_alias = _resolve_item(group_col)
        if group_alias is None:
            raise PreprocessError(f"Group column '{group_col}' was not found in the dataset.")
        groups = {
            str(row.get(group_alias))
            for row in dict_rows
            if row.get(group_alias) is not None and str(row.get(group_alias)) != ""
        }
        if len(groups) != 2:
            raise PreprocessError(
                f"Template 'ttest_two_group' requires exactly 2 groups, but found {len(groups)}: {', '.join(sorted(groups))}."
            )

    elif template_hint == "prepost_scale_study":
        pre_prefix = spec.get("pre_prefix", "pre_")
        post_prefix = spec.get("post_prefix", "post_")
        id_column = spec.get("id_column")
        min_scale = spec.get("min_scale", 1)
        max_scale = spec.get("max_scale", 5)

        pre_aliases = [a for a in aliases if a.startswith(pre_prefix)]
        post_aliases = [a for a in aliases if a.startswith(post_prefix)]
        if not pre_aliases:
            raise PreprocessError(
                f"Template 'prepost_scale_study' requires at least one column with prefix '{pre_prefix}'."
            )
        if not post_aliases:
            raise PreprocessError(
                f"Template 'prepost_scale_study' requires at least one column with prefix '{post_prefix}'."
            )

        if id_column:
            id_alias = _resolve_item(id_column)
            if id_alias is None:
                raise PreprocessError(f"id_column '{id_column}' was not found in the dataset.")
            seen_ids: set[Any] = set()
            duplicate_ids: set[Any] = set()
            for r in dict_rows:
                val = r.get(id_alias)
                if val is None:
                    continue
                if val in seen_ids:
                    duplicate_ids.add(val)
                else:
                    seen_ids.add(val)
            if duplicate_ids:
                sample = ", ".join(str(v) for v in list(duplicate_ids)[:5])
                raise PreprocessError(f"Duplicate id values detected in '{id_column}': {sample}")

        _validate_integer_scale(dict_rows, pre_aliases + post_aliases, manifest, min_scale, max_scale)

    elif template_hint == "cross_sectional_survey":
        min_scale = spec.get("min_scale", 1)
        max_scale = spec.get("max_scale", 5)
        scale_aliases = [a for a in aliases if re.match(r"^q\d+$", a, re.IGNORECASE)]
        if not scale_aliases:
            raise PreprocessError(
                "Template 'cross_sectional_survey' requires at least one column matching 'q01', 'q02', etc."
            )

        user_id_alias = _resolve_item("user_id")
        if user_id_alias is not None:
            seen_ids: set[Any] = set()
            duplicate_ids: set[Any] = set()
            for r in dict_rows:
                val = r.get(user_id_alias)
                if val is None:
                    continue
                if val in seen_ids:
                    duplicate_ids.add(val)
                else:
                    seen_ids.add(val)
            if duplicate_ids:
                sample = ", ".join(str(v) for v in list(duplicate_ids)[:5])
                raise PreprocessError(f"Duplicate user_id values detected: {sample}")

        _validate_integer_scale(dict_rows, scale_aliases, manifest, min_scale, max_scale)

    elif template_hint == "regression_study":
        dep_candidates = [spec.get("dependent_var"), "y", "dep", "dependent"]
        dep_alias: str | None = None
        for cand in dep_candidates:
            if cand:
                resolved = _resolve_item(cand)
                if resolved is not None:
                    dep_alias = resolved
                    break
        if dep_alias is None:
            raise PreprocessError(
                "Template 'regression_study' requires a dependent variable column (default 'y' or specify 'dependent_var')."
            )

        # Verify dependent variable is numeric
        non_numeric = 0
        for i, r in enumerate(dict_rows):
            val = r.get(dep_alias)
            if val is None or val == "":
                continue
            try:
                float(val)
            except (ValueError, TypeError):
                non_numeric += 1
        if non_numeric > len(dict_rows) // 2:
            raise PreprocessError(
                f"Dependent variable '{manifest[dep_alias]}' must be numeric for regression analysis."
            )

        # Require at least one predictor (any column besides the dependent)
        predictor_aliases = [a for a in aliases if a != dep_alias]
        if not predictor_aliases:
            raise PreprocessError(
                "Template 'regression_study' requires at least one predictor column in addition to the dependent variable."
            )


def _validate_integer_scale(
    dict_rows: list[dict[str, Any]],
    scale_aliases: list[str],
    manifest: dict[str, str],
    min_scale: int,
    max_scale: int,
) -> None:
    for alias in scale_aliases:
        for i, row in enumerate(dict_rows):
            val = row.get(alias)
            if val is None or val == "":
                continue
            try:
                num = float(val)
                if not num.is_integer():
                    raise PreprocessError(
                        f"Column '{manifest[alias]}' (row {i + 1}) contains a non-integer value: {val}. "
                        f"Scale data must be integers between {min_scale} and {max_scale}."
                    )
                num = int(num)
                if num < min_scale or num > max_scale:
                    raise PreprocessError(
                        f"Column '{manifest[alias]}' (row {i + 1}) value {num} is out of range. "
                        f"Expected integers between {min_scale} and {max_scale}."
                    )
            except (ValueError, TypeError):
                raise PreprocessError(
                    f"Column '{manifest[alias]}' (row {i + 1}) contains an invalid value: {val}. "
                    f"Scale data must be integers between {min_scale} and {max_scale}."
                )


def preprocess_data(
    data_path: Path,
    spec: dict[str, Any],
    output_dir: Path | None = None,
) -> tuple[Path, dict[str, str], dict[str, Any], dict[str, Any]]:
    suffix = data_path.suffix.lower()
    supported_text = {".csv", ".tsv", ".txt"}
    supported_excel = {".xlsx", ".xlsm"}
    if suffix not in supported_text | supported_excel:
        raise PreprocessError(
            f"Unsupported data file type '{suffix}'. Supported formats: {', '.join(sorted(supported_text | supported_excel))}."
        )
    headers: list[str] = []
    rows: list[list[Any]] = []

    if suffix in supported_excel:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            raise PreprocessError("openpyxl is required to process Excel files")

        workbook = openpyxl.load_workbook(data_path, read_only=True, data_only=True)
        sheet = workbook.active
        if spec.get("sheet") and spec["sheet"] in workbook.sheetnames:
            sheet = workbook[spec["sheet"]]

        row_iter = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(row_iter)
        except StopIteration:
            raw_headers = ()

        for h in raw_headers:
            headers.append(str(h).strip() if h is not None else "")

        for row in row_iter:
            rows.append(list(row))
        workbook.close()
    else:
        delimiter = "\t" if suffix == ".tsv" else ","
        try:
            dialect = csv.Sniffer().sniff(data_path.read_text(encoding="utf-8-sig")[:4096], delimiters=",;\t|")
            delimiter = dialect.delimiter
        except Exception:
            pass

        with data_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            try:
                raw_headers = next(reader)
            except StopIteration:
                raw_headers = []

            for h in raw_headers:
                headers.append(str(h).strip())

            for row in reader:
                rows.append(row)

    seen: set[str] = set()
    manifest: dict[str, str] = {}
    aliases: list[str] = []
    for i, h in enumerate(headers):
        alias = safe_alias(h, i, seen)
        aliases.append(alias)
        manifest[alias] = h

    dict_rows = []
    for row in rows:
        d = {}
        for i, a in enumerate(aliases):
            d[a] = row[i] if i < len(row) else None
        dict_rows.append(d)

    is_preset = spec.get("request_kind") == "preset"
    updated_spec = dict(spec)
    alias_seen = set(aliases)

    def _resolve_item(item_name: str) -> str | None:
        """Resolve an item name to its alias. Matches by alias key first, then by original header value."""
        if item_name in manifest:
            return item_name
        for k, v in manifest.items():
            if v == item_name:
                return k
        return None

    if is_preset and "preset" in spec:
        preset = spec["preset"]
        max_scale = preset.get("max_scale", 5)
        rev_items = preset.get("reverse_items", [])
        subscales = preset.get("subscales", {})

        if preset.get("name") == "prepost_scale_study":
            pre_prefix = preset.get("pre_prefix", "pre_")
            post_prefix = preset.get("post_prefix", "post_")
            id_column = preset.get("id_column")
            group_col = preset.get("group_column")
            cluster_col = preset.get("cluster_column")

            if id_column:
                id_alias = _resolve_item(id_column)
                if id_alias is None:
                    raise PreprocessError(f"id_column '{id_column}' was not found in the dataset.")
                seen_ids: set[Any] = set()
                duplicate_ids: set[Any] = set()
                for r in dict_rows:
                    val = r.get(id_alias)
                    if val is None:
                        continue
                    if val in seen_ids:
                        duplicate_ids.add(val)
                    else:
                        seen_ids.add(val)
                if duplicate_ids:
                    sample = ", ".join(str(v) for v in list(duplicate_ids)[:5])
                    raise PreprocessError(f"Duplicate id values detected in '{id_column}': {sample}")

            reverse_aliases: dict[str, str] = {}

            def resolve_prefixed(item: str, prefix: str) -> str | None:
                candidate = item if item.startswith(prefix) else f"{prefix}{item}"
                return _resolve_item(candidate)

            for rev_item_name in rev_items:
                for prefix in (pre_prefix, post_prefix):
                    resolved = resolve_prefixed(rev_item_name, prefix)
                    if resolved is None or resolved in reverse_aliases:
                        continue
                    rev_alias = safe_alias(f"{resolved}_rev", 0, alias_seen)
                    manifest[rev_alias] = f"Reversed({manifest[resolved]})"
                    aliases.append(rev_alias)
                    alias_seen.add(rev_alias)
                    reverse_aliases[resolved] = rev_alias
                    for r in dict_rows:
                        val = r.get(resolved)
                        try:
                            r[rev_alias] = (max_scale + 1) - float(val) if val is not None else None
                        except (ValueError, TypeError):
                            r[rev_alias] = None

            pre_aliases: list[str] = []
            post_aliases: list[str] = []
            delta_aliases: list[str] = []
            pair_specs: list[dict[str, str]] = []

            for sub_name, items in subscales.items():
                pre_items: list[str] = []
                post_items: list[str] = []
                for item in items:
                    pre_item = resolve_prefixed(item, pre_prefix)
                    if pre_item is not None:
                        pre_items.append(reverse_aliases.get(pre_item, pre_item))
                    post_item = resolve_prefixed(item, post_prefix)
                    if post_item is not None:
                        post_items.append(reverse_aliases.get(post_item, post_item))

                if not pre_items and not post_items:
                    continue

                pre_alias = safe_alias(f"{pre_prefix}{sub_name}", 0, alias_seen)
                post_alias = safe_alias(f"{post_prefix}{sub_name}", 0, alias_seen)
                delta_alias = safe_alias(f"delta_{sub_name}", 0, alias_seen)
                aliases.extend([pre_alias, post_alias, delta_alias])
                manifest[pre_alias] = f"Subscale({sub_name}) Pre"
                manifest[post_alias] = f"Subscale({sub_name}) Post"
                manifest[delta_alias] = f"Subscale({sub_name}) Delta"

                for r in dict_rows:
                    pre_vals: list[float] = []
                    post_vals: list[float] = []
                    for ik in pre_items:
                        try:
                            if r.get(ik) is not None:
                                pre_vals.append(float(r[ik]))
                        except (ValueError, TypeError):
                            pass
                    for ik in post_items:
                        try:
                            if r.get(ik) is not None:
                                post_vals.append(float(r[ik]))
                        except (ValueError, TypeError):
                            pass
                    pre_value = sum(pre_vals) / len(pre_vals) if pre_vals else None
                    post_value = sum(post_vals) / len(post_vals) if post_vals else None
                    r[pre_alias] = pre_value
                    r[post_alias] = post_value
                    if pre_value is not None and post_value is not None:
                        r[delta_alias] = post_value - pre_value
                    else:
                        r[delta_alias] = None

                pre_aliases.append(pre_alias)
                post_aliases.append(post_alias)
                delta_aliases.append(delta_alias)
                if pre_items and post_items:
                    pair_specs.append({"i1": pre_alias, "i2": post_alias})

            prepost_analyses: list[dict[str, Any]] = []
            desc_vars: dict[str, Any] = {"vars": pre_aliases + post_aliases + delta_aliases}
            split_cols = [col for col in (group_col, cluster_col) if col]
            if split_cols:
                desc_vars["splitBy"] = split_cols
            prepost_analyses.append({"analysis_type": "descriptives", "variables": desc_vars})

            if pair_specs:
                prepost_analyses.append({"analysis_type": "ttestPS", "variables": {"pairs": pair_specs}})

            if group_col and delta_aliases:
                prepost_analyses.append({
                    "analysis_type": "ttestIS",
                    "variables": {"vars": delta_aliases, "group": group_col},
                })

            updated_spec["analyses"] = prepost_analyses
        else:
            for rev_item_name in rev_items:
                resolved = _resolve_item(rev_item_name)
                if resolved is None:
                    continue
                rev_alias = f"{resolved}_rev"
                if rev_alias in manifest:
                    continue
                manifest[rev_alias] = f"Reversed({manifest[resolved]})"
                aliases.append(rev_alias)
                alias_seen.add(rev_alias)
                for r in dict_rows:
                    val = r.get(resolved)
                    try:
                        r[rev_alias] = (max_scale + 1) - float(val) if val is not None else None
                    except (ValueError, TypeError):
                        r[rev_alias] = None

            subscale_aliases: dict[str, str] = {}
            for sub_name, items in subscales.items():
                sub_alias = safe_alias(sub_name, 0, alias_seen)
                aliases.append(sub_alias)
                manifest[sub_alias] = f"Subscale({sub_name})"
                subscale_aliases[sub_name] = sub_alias

                item_keys: list[str] = []
                for item in items:
                    resolved = _resolve_item(item)
                    if resolved is None:
                        continue
                    if f"{resolved}_rev" in manifest:
                        item_keys.append(f"{resolved}_rev")
                    else:
                        item_keys.append(resolved)

                for r in dict_rows:
                    vals = []
                    for ik in item_keys:
                        try:
                            if r.get(ik) is not None:
                                vals.append(float(r[ik]))
                        except (ValueError, TypeError):
                            pass
                    r[sub_alias] = sum(vals) / len(vals) if vals else None

    _validate_template(dict_rows, aliases, manifest, updated_spec)

    if "analyses" in updated_spec:
        inv_manifest = {v: k for k, v in manifest.items()}
        for an in updated_spec["analyses"]:
            if "variables" in an:
                for role, binding in an["variables"].items():
                    if isinstance(binding, list):
                        if role == "pairs":
                            new_pairs = []
                            for p in binding:
                                if not isinstance(p, dict):
                                    continue
                                new_p = {}
                                if "i1" in p:
                                    new_p["i1"] = inv_manifest.get(p["i1"], p["i1"])
                                if "i2" in p:
                                    new_p["i2"] = inv_manifest.get(p["i2"], p["i2"])
                                if new_p:
                                    new_pairs.append(new_p)
                            an["variables"][role] = new_pairs
                        else:
                            an["variables"][role] = [inv_manifest.get(v, v) for v in binding]
                    elif isinstance(binding, str):
                        an["variables"][role] = inv_manifest.get(binding, binding)

    if "measure_overrides" in updated_spec:
        inv_manifest = {v: k for k, v in manifest.items()}
        new_mo = {}
        for k, v in updated_spec["measure_overrides"].items():
            new_mo[inv_manifest.get(k, k)] = v
        updated_spec["measure_overrides"] = new_mo

    sidecar_dir = Path(output_dir) if output_dir is not None else None
    if sidecar_dir is None:
        sidecar_dir = Path(tempfile.mkdtemp(prefix="jamovi-preprocess-"))
    else:
        sidecar_dir.mkdir(parents=True, exist_ok=True)

    out_csv = sidecar_dir / "analysis_ready.csv"
    with out_csv.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=aliases)
        writer.writeheader()
        writer.writerows(dict_rows)

    out_manifest = sidecar_dir / "column_manifest.json"
    with out_manifest.open("w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

    sidecar_info = {
        "dir": sidecar_dir,
        "analysis_ready": out_csv,
        "column_manifest": out_manifest,
    }

    return out_csv, manifest, updated_spec, sidecar_info
